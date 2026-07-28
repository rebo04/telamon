#!/usr/bin/env python3
"""
Genera src/tis-f0124.js a partir del documento controlado
"TIS-F0124 - Código de Defectos (S).xlsx".

Uso:
    python3 tools/gen-catalog.py <ruta-al-xlsx>

Los NOMBRES y CÓDIGOS salen íntegros del documento: no se editan a mano.
Las CLASIFICACIONES (clase de elemento, clase de tipo) y los ANCLAJES de
severidad son criterio de ingeniería y viven en este script — ver README.

Al liberarse una nueva revisión del documento:
  1. Reemplazar el .xlsx.
  2. Revisar que los elementos/tipos nuevos estén clasificados abajo
     (el script aborta si encuentra alguno sin clasificar).
  3. Re-ejecutar. Subir TIS_F0124_REV y TIS_F0124_FECHA.
  4. NUNCA renumerar ni reasignar un código existente: se deprecan.
"""

import sys
import re
import unicodedata
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "src" / "tis-f0124.js"

# ── Metadatos del documento controlado ───────────────────────────────────
REV = 9
FECHA = "2024-09-17"

# ── Clasificación de ELEMENTOS (criterio de ingeniería) ──────────────────
# ELEC  Portador de corriente o responsable de retención de terminal.
# IDENT Identificación y trazabilidad.
# MEC   Ensamble mecánico / sujeción / dimensional.
# PROT  Protección y acabado (encintado, tubos, mangas).
# PROC  Equipo o insumo de proceso — no es parte del producto.
ELEM_CLASS = {
    "ELEC": ["AI", "BC", "CA", "CM", "CS", "CV", "CI", "CP", "CN", "CL", "CPA",
             "DE", "EM", "EST", "FE", "HI", "LT", "PS", "PC", "PA", "SO", "SW",
             "TE", "AL", "PI", "TS", "BT", "PIG", "RES", "ES"],
    "IDENT": ["CB", "ET", "MT", "CE", "RAS"],
    "MEC": ["AM", "AR", "ARA", "BA", "BO", "BU", "CAM", "CAN", "CO", "CR", "CC",
            "RO", "REM", "SE", "TA", "TO", "VE", "RA", "TU", "DI", "RE", "GRA",
            "FLS"],
    "PROT": ["EN", "ENM", "EE", "SC", "TC", "TP", "TR", "TT", "MAN"],
    "PROC": ["EQ", "MA", "BR", "MU"],
}

# ── Clasificación de TIPOS (criterio de ingeniería) ──────────────────────
# FUNC  Compromete la función eléctrica o la integridad del elemento.
# IDENT Impide o falsea la identificación.
# ENS   Desviación de ensamble o dimensional.
# COSM  Acabado / estético.
# DISP  NO ES DEFECTO: razón de disposición de material o error de prueba.
#       Se excluye de los KPI de defecto y no lleva severidad.
TYPE_CLASS = {
    "FUNC": [1, 4, 5, 11, 12, 13, 14, 15, 19, 27, 28, 32, 33, 35, 36, 37, 38,
             46, 47, 48, 49, 50, 56, 57, 60, 66, 72],
    "IDENT": [25, 30, 68],
    "ENS": [3, 7, 8, 9, 10, 16, 17, 21, 22, 23, 24, 26, 29, 31, 34, 40, 41, 42,
            55, 58, 61, 63, 64, 65, 67, 69, 70],
    "COSM": [2, 6, 18, 39, 59, 62, 71],
    "DISP": [20, 43, 44, 45, 51, 52, 53, 54],
}

# ── Matriz clase-elemento × clase-tipo → severidad ───────────────────────
# Se aplica SÓLO cuando ningún anclaje documental cubre el par.
MATRIX = {
    "ELEC":  {"FUNC": "A", "IDENT": "B", "ENS": "B", "COSM": "C"},
    "IDENT": {"FUNC": "A", "IDENT": "A", "ENS": "B", "COSM": "C"},
    "MEC":   {"FUNC": "B", "IDENT": "B", "ENS": "B", "COSM": "C"},
    "PROT":  {"FUNC": "C", "IDENT": "C", "ENS": "C", "COSM": "C"},
    "PROC":  {"FUNC": "B", "IDENT": "B", "ENS": "B", "COSM": "C"},
}

DEFAULT_SEV = "B"

# ── Anclajes documentales (Sheet1 — SEVERIDAD / CRITERIO) ────────────────
# Cada entrada cita el texto del documento que la justifica. Ganan sobre la
# matriz. Formato: (elementos, tipos, severidad, cita)
DOC_ANCHORS = [
    (["PC", "PA"], [14, 11, 50, 22, 19, 32], "A",
     "A · Falla de prensado (aislante dentro del prensado, hilos cortados o "
     "fuera del prensado, no hilos dentro del prensado)"),
    (["HI"], [11, 50, 22, 14], "A",
     "A · Falla de prensado (hilos cortados o fuera del prensado, no hilos "
     "dentro del prensado)"),
    (["AI"], [14], "A",
     "A · Falla de prensado (aislante dentro del prensado)"),
    (["TE"], [32, 28], "A",
     "A · Terminal no asentada / terminal jalada"),
    (["TE"], [15, 57], "A",
     "A · Terminales desalineadas / dobladas"),
    (["CB"], [25, 22, 26], "A",
     "A · Barcode ilegible / falta de barcode / barcode no corresponde al "
     "arnés o NP"),
    (["DI", "CO", "CR", "CC", "RA", "TP"], [24], "B",
     "B · Problemas dimensionales (clip, ramales, pvc fuera de tolerancia)"),
    (["CO", "CR", "CC", "EN", "CS", "SE"], [22], "B",
     "B · Falta de componentes (clip faltante, encintado, candados "
     "secundarios faltantes)"),
    (["CM", "CS"], [1], "B",
     "B · Falta de componentes (candados abiertos)"),
    (["CO", "CR", "CC"], [26], "B",
     "B · Clip equivocado"),
    (["EN"], [6, 18, 61], "C",
     "C · Encintado en bandera, encintado pobre, exceso de cinta"),
    (["SC"], [18], "C",
     "C · Exceso de corbata en clip"),
    (["CO", "CR", "CC"], [42], "C",
     "C · Clip deforme (se puede ensamblar)"),
    (["TP", "TC", "TR"], [11, 16], "C",
     "C · PVC, corrugado, tubo de calor, mal cortado (sesgado)"),
]

# ── Migración del catálogo legado ────────────────────────────────────────
# Sólo se auto-mapea cuando la etiqueta vieja determina AMBOS ejes a una
# única entrada. Lo ambiguo migra a SIN-CLASIFICAR conservando el texto.
LEGACY_MAP = {
    "Terminal no insertada (Push-back)":      ("TE", 32),
    "Cables invertidos (Miswire)":            ("CA", 27),
    "Circuito abierto / Sin continuidad":     ("CN", 1),
    "Cortocircuito":                          ("CN", 12),
    "Terminal dañada / Deformada":            ("TE", 56),
}
LEGACY_AMBIGUOUS = {
    "Daño en aislamiento / Cobre expuesto":
        "Dos modos distintos: AI-56 (aislante dañado) vs HI-19 (hilos expuestos)",
    "Falta de componente (Clip, Sello, TPA)":
        "La etiqueta nombra tres elementos candidatos: CO/CR/CC, SE y CPA",
    "Conector dañado / Roto":
        "Tipo determinado (36) pero elemento no: CP, CI o CL",
    "Ruteo incorrecto / Longitud":
        "AR-70 (mal ruteado) vs DI-24 (fuera de tolerancia)",
    "Encintado defectuoso / Faltante":
        "EN-61 (incompleto) vs EN-22 (faltante)",
}


def js(s):
    return "'" + (str(s)
                  .replace("\\", "\\\\")
                  .replace("'", "\\'")
                  .replace("\r\n", "\\n")
                  .replace("\r", "\\n")
                  .replace("\n", "\\n")) + "'"


def read_doc(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Codigo de defectos "]
    elements, types = [], []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        name, code, tname, tnum = (c.value for c in row)
        if code and name:
            elements.append((str(code).strip(), str(name).strip()))
        if tnum is not None and tname:
            types.append((int(tnum), str(tname).strip()))
    sev = []
    for row in wb["Sheet1"].iter_rows(min_row=2, max_row=4, max_col=2):
        letra, criterio = row[0].value, row[1].value
        if letra and criterio:
            sev.append((str(letra).strip(), str(criterio).strip()))
    return elements, types, sev


def invert(mapping, label):
    out = {}
    for cls, keys in mapping.items():
        for k in keys:
            if k in out:
                sys.exit(f"ERROR: {label} {k!r} clasificado dos veces "
                         f"({out[k]} y {cls})")
            out[k] = cls
    return out


def fold(s):
    """Normaliza para búsqueda: minúsculas sin acentos."""
    return "".join(c for c in unicodedata.normalize("NFD", s.lower())
                   if unicodedata.category(c) != "Mn")


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    elements, types, sev_table = read_doc(src)

    ec = invert(ELEM_CLASS, "elemento")
    tc = invert(TYPE_CLASS, "tipo")

    missing_e = [c for c, _ in elements if c not in ec]
    missing_t = [n for n, _ in types if n not in tc]
    extra_e = [c for c in ec if c not in {c for c, _ in elements}]
    extra_t = [n for n in tc if n not in {n for n, _ in types}]
    if missing_e or missing_t or extra_e or extra_t:
        sys.exit(f"ERROR: clasificación desincronizada del documento.\n"
                 f"  elementos sin clasificar: {missing_e}\n"
                 f"  tipos sin clasificar:     {missing_t}\n"
                 f"  elementos inexistentes:   {extra_e}\n"
                 f"  tipos inexistentes:       {extra_t}")

    # Anclajes → mapa plano "EL-N" → [sev, cita]
    anchors = {}
    for els, tps, s, cite in DOC_ANCHORS:
        for e in els:
            if e not in ec:
                sys.exit(f"ERROR: anclaje sobre elemento inexistente {e!r}")
            for t in tps:
                if t not in tc:
                    sys.exit(f"ERROR: anclaje sobre tipo inexistente {t!r}")
                key = f"{e}-{t}"
                if key in anchors and anchors[key][0] != s:
                    sys.exit(f"ERROR: anclajes en conflicto para {key}: "
                             f"{anchors[key][0]} y {s}")
                anchors[key] = (s, cite)

    for label, (e, t) in LEGACY_MAP.items():
        if f"{e}-{t}" and (e not in ec or t not in tc):
            sys.exit(f"ERROR: mapeo legado inválido para {label!r}")

    L = []
    w = L.append
    w("// ─────────────────────────────────────────────────────────────────────")
    w("// ARCHIVO GENERADO — NO EDITAR A MANO.")
    w("// Fuente: TIS-F0124 - Código de Defectos (S).xlsx")
    w("// Regenerar: python3 tools/gen-catalog.py <ruta-al-xlsx>")
    w("// ─────────────────────────────────────────────────────────────────────")
    w("")
    w(f"export const TIS_F0124_FORM  = 'TIS-F0124';")
    w(f"export const TIS_F0124_REV   = {REV};")
    w(f"export const TIS_F0124_FECHA = {js(FECHA)};")
    w(f"export const TIS_F0124_SELLO = "
      f"{js(f'TIS-F0124 Rev.{REV} ({FECHA})')};")
    w("")

    w("// ── ELEMENTOS (columna A/B del documento) ──────────────────────────")
    w("export const ELEMENTOS = [")
    for code, name in elements:
        w(f"  {{ code: {js(code)}, nombre: {js(name)}, "
          f"clase: {js(ec[code])}, buscar: {js(fold(code + ' ' + name))} }},")
    w("];")
    w("")

    w("// ── TIPOS DE DEFECTO (columna C/D del documento) ───────────────────")
    w("export const TIPOS = [")
    for num, name in types:
        w(f"  {{ num: {num}, nombre: {js(name)}, "
          f"clase: {js(tc[num])}, buscar: {js(fold(str(num) + ' ' + name))} }},")
    w("];")
    w("")

    w("// ── SEVERIDAD (hoja Sheet1 del documento) ──────────────────────────")
    w("export const SEVERIDADES = {")
    for letra, criterio in sev_table:
        w(f"  {js(letra)}: {js(criterio)},")
    w("};")
    w("")

    w("// ── Anclajes documentales: ganan sobre la matriz ───────────────────")
    w("export const ANCLAJES_DOC = {")
    for key in sorted(anchors, key=lambda k: (k.split("-")[0], int(k.split("-")[1]))):
        s, cite = anchors[key]
        w(f"  {js(key)}: {{ sev: {js(s)}, cita: {js(cite)} }},")
    w("};")
    w("")

    w("// ── Matriz clase-elemento × clase-tipo ─────────────────────────────")
    w("export const MATRIZ_SEVERIDAD = {")
    for e, row in MATRIX.items():
        cells = ", ".join(f"{k}: {js(v)}" for k, v in row.items())
        w(f"  {e}: {{ {cells} }},")
    w("};")
    w("")
    w(f"export const SEVERIDAD_DEFAULT = {js(DEFAULT_SEV)};")
    w("")

    w("// ── Migración del catálogo legado ──────────────────────────────────")
    w("export const LEGADO_DETERMINISTA = {")
    for label, (e, t) in LEGACY_MAP.items():
        w(f"  {js(label)}: {{ el: {js(e)}, tipo: {t} }},")
    w("};")
    w("")
    w("export const LEGADO_AMBIGUO = {")
    for label, motivo in LEGACY_AMBIGUOUS.items():
        w(f"  {js(label)}: {js(motivo)},")
    w("};")
    w("")

    OUT.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"✓ {OUT.relative_to(REPO)}")
    print(f"  {len(elements)} elementos · {len(types)} tipos · "
          f"{len(anchors)} anclajes documentales")
    print(f"  {len(LEGACY_MAP)} mapeos deterministas · "
          f"{len(LEGACY_AMBIGUOUS)} etiquetas ambiguas")


if __name__ == "__main__":
    main()
